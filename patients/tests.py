import json
from datetime import datetime, timezone as datetime_timezone
from io import BytesIO
from pathlib import Path
from unittest.mock import Mock, patch

from django.contrib import admin
from django.contrib.auth import get_user_model
from django.contrib.messages import get_messages
from django.db import DatabaseError, IntegrityError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase.pdfmetrics import Font

from .admin import (
    PatientAdmin,
    PatientAdminForm,
    build_patients_excel,
    build_patients_pdf,
    SMSMessageLogAdmin,
    SMSMessageLogInline,
    format_sms_response,
)
from .datetime import (
    format_tehran_jalali,
    format_tehran_jalali_input,
    parse_tehran_jalali_datetime,
    to_english_digits,
)
from .forms import (
    DUPLICATE_MOBILE_ERROR,
    DUPLICATE_NATIONAL_CODE_ERROR,
    PatientRegistrationForm,
)
from .views import (
    COMMUNITY_BASE_COUNT,
    CANONICAL_URL,
    FORM_ERROR_MESSAGE,
    SHARE_DESCRIPTION,
    SHARE_IMAGE_PATH,
    SHARE_TITLE,
    SITE_LOGO_PATH,
    SITE_NAME,
    SUCCESS_MESSAGE,
)
from .models import SMSMessageLog, Patient, VisitEvent
from .sms import (
    KavenegarSMSDeliveryError,
    KavenegarSMSConfigurationError,
    build_patient_name_token,
    send_done_sms,
    send_register_sms,
)


def register_test_pdf_fonts():
    registered_fonts = set(pdfmetrics.getRegisteredFontNames())
    if "DejaVuSans" not in registered_fonts:
        pdfmetrics.registerFont(Font("DejaVuSans", "Helvetica", "WinAnsiEncoding"))
    if "DejaVuSans-Bold" not in registered_fonts:
        pdfmetrics.registerFont(
            Font("DejaVuSans-Bold", "Helvetica-Bold", "WinAnsiEncoding")
        )
    registerFontFamily(
        "DejaVuSans",
        normal="DejaVuSans",
        bold="DejaVuSans-Bold",
        italic="DejaVuSans",
        boldItalic="DejaVuSans-Bold",
    )


class PatientModelTests(TestCase):
    def test_mobile_field_is_unique(self):
        self.assertTrue(Patient._meta.get_field("mobile").unique)


class PersianDateTimeFormatTests(TestCase):
    def test_format_tehran_jalali_converts_utc_to_tehran_and_jalali(self):
        value = datetime(2026, 3, 20, 20, 45, 10, tzinfo=datetime_timezone.utc)

        self.assertEqual(format_tehran_jalali(value), "۱۴۰۵/۰۱/۰۱ ۰۰:۱۵:۱۰")

    def test_format_tehran_jalali_handles_naive_datetime(self):
        value = datetime(2026, 3, 20, 20, 45, 10)

        self.assertEqual(format_tehran_jalali(value), "۱۴۰۴/۱۲/۲۹ ۲۰:۴۵:۱۰")

    def test_jalali_input_format_and_parser_accept_persian_digits(self):
        value = datetime(2026, 6, 26, 5, 30, tzinfo=datetime_timezone.utc)

        self.assertEqual(format_tehran_jalali_input(value), "۱۴۰۵/۰۴/۰۵ ۰۹:۰۰")
        self.assertEqual(to_english_digits("۱۴۰۵/۰۴/۰۵ ۰۹:۰۰"), "1405/04/05 09:00")
        self.assertEqual(
            parse_tehran_jalali_datetime("۱۴۰۵/۰۴/۰۵ ۰۹:۰۰").astimezone(
                datetime_timezone.utc
            ),
            value,
        )

    def test_jalali_input_parser_accepts_dash_and_defaults_seconds(self):
        parsed = parse_tehran_jalali_datetime("1405-04-05 09:00")

        self.assertEqual(parsed.second, 0)
        self.assertEqual(format_tehran_jalali(parsed), "۱۴۰۵/۰۴/۰۵ ۰۹:۰۰:۰۰")


class PatientRegistrationFormTests(TestCase):
    def test_valid_registration_form(self):
        form = PatientRegistrationForm(
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "09123456789",
            }
        )

        self.assertTrue(form.is_valid())

    def test_national_code_must_be_exactly_ten_digits(self):
        form = PatientRegistrationForm(
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "123456789",
                "mobile": "09123456789",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertEqual(form.errors["national_code"], ["کد ملی را ۱۰ رقمی وارد کنید."])

    def test_national_code_must_contain_only_digits(self):
        form = PatientRegistrationForm(
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "123456789a",
                "mobile": "09123456789",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertEqual(
            form.errors["national_code"], ["کد ملی باید فقط شامل عدد باشد."]
        )

    def test_national_code_must_be_unique(self):
        Patient.objects.create(
            first_name="Existing",
            last_name="Patient",
            national_code="1234567890",
            mobile="09111111111",
        )
        form = PatientRegistrationForm(
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "09123456789",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertEqual(form.errors["national_code"], [DUPLICATE_NATIONAL_CODE_ERROR])

    def test_mobile_must_be_exactly_eleven_digits(self):
        form = PatientRegistrationForm(
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "0912345678",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertEqual(
            form.errors["mobile"],
            ["شماره را ۱۱ رقمی وارد کنید."],
        )

    def test_mobile_must_contain_only_digits(self):
        form = PatientRegistrationForm(
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "0912345678a",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertEqual(
            form.errors["mobile"],
            ["فقط عدد وارد کنید."],
        )

    def test_mobile_must_start_with_09(self):
        form = PatientRegistrationForm(
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "08123456789",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertEqual(
            form.errors["mobile"],
            ["شماره موبایل با 09 شروع شود."],
        )

    def test_mobile_must_be_unique(self):
        Patient.objects.create(
            first_name="Existing",
            last_name="Patient",
            mobile="09123456789",
            national_code="1111111111",
        )
        form = PatientRegistrationForm(
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "09123456789",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertEqual(
            form.errors["mobile"],
            [DUPLICATE_MOBILE_ERROR],
        )

    def test_required_field_messages_are_persian(self):
        form = PatientRegistrationForm(data={})

        self.assertFalse(form.is_valid())
        self.assertEqual(form.errors["first_name"], ["نام را وارد کنید."])
        self.assertEqual(form.errors["last_name"], ["نام خانوادگی را وارد کنید."])
        self.assertEqual(form.errors["national_code"], ["کد ملی را وارد کنید."])
        self.assertEqual(form.errors["mobile"], ["شماره موبایل را وارد کنید."])


class KavenegarRegisterSMSTests(TestCase):
    def test_patient_name_token_replaces_spaces_inside_first_and_last_name(self):
        patient = Patient(first_name="علی رضا", last_name="کاوه نگار")

        self.assertEqual(build_patient_name_token(patient), "علی_رضا_کاوه_نگار")

    @override_settings(
        KAVENEGAR_API_KEY="test-api-key",
        KAVENEGAR_REGISTER_TEMPLATE="register-template",
        KAVENEGAR_REQUEST_TIMEOUT_SECONDS=7,
    )
    def test_send_register_sms_uses_configured_template_as_template_token(self):
        kavenegar_response = Mock()
        kavenegar_response.json.return_value = {
            "return": {"status": 200},
            "entries": [{"messageid": 123}],
        }

        with patch(
            "patients.sms.requests.post", return_value=kavenegar_response
        ) as post:
            result = send_register_sms("09123456789", "Ali_Ahmadi")

        self.assertEqual(result, [{"messageid": 123}])
        post.assert_called_once()
        self.assertEqual(
            post.call_args.args[0],
            "https://api.kavenegar.com/v1/test-api-key/verify/lookup.json",
        )
        self.assertEqual(
            post.call_args.kwargs["data"],
            {
                "receptor": "09123456789",
                "template": "register-template",
                "token": "Ali_Ahmadi",
            },
        )
        self.assertEqual(post.call_args.kwargs["timeout"], 7)

    @override_settings(
        KAVENEGAR_API_KEY="test-api-key",
        KAVENEGAR_DONE_TEMPLATE="done-template",
    )
    def test_send_done_sms_uses_configured_done_template(self):
        kavenegar_response = Mock()
        kavenegar_response.json.return_value = {
            "return": {"status": 200},
            "entries": [{"messageid": 456}],
        }

        with patch(
            "patients.sms.requests.post", return_value=kavenegar_response
        ) as post:
            result = send_done_sms("09123456789", "Ali_Ahmadi")

        self.assertEqual(result, [{"messageid": 456}])
        self.assertEqual(
            post.call_args.kwargs["data"],
            {
                "receptor": "09123456789",
                "template": "done-template",
                "token": "Ali_Ahmadi",
            },
        )

    @override_settings(KAVENEGAR_API_KEY="test-api-key", KAVENEGAR_REGISTER_TEMPLATE="")
    def test_send_register_sms_requires_configured_template(self):
        with self.assertRaises(KavenegarSMSConfigurationError):
            send_register_sms("09123456789", "Ali_Ahmadi")

    @override_settings(KAVENEGAR_API_KEY="")
    def test_send_register_sms_requires_api_key(self):
        with self.assertRaises(KavenegarSMSConfigurationError):
            send_register_sms("09123456789", "Ali_Ahmadi")

    @override_settings(
        KAVENEGAR_API_KEY="test-api-key",
        KAVENEGAR_REGISTER_TEMPLATE="register-template",
    )
    def test_send_register_sms_raises_delivery_error_on_timeout(self):
        with patch(
            "patients.sms.requests.post",
            side_effect=__import__("requests").exceptions.Timeout,
        ):
            with self.assertRaises(KavenegarSMSDeliveryError):
                send_register_sms("09123456789", "Ali_Ahmadi")

    @override_settings(
        KAVENEGAR_API_KEY="test-api-key",
        KAVENEGAR_REGISTER_TEMPLATE="register-template",
        SMS_SEND_ASYNC=False,
    )
    def test_patient_creation_signal_sends_register_sms_after_commit(self):
        with patch("patients.sms_tasks.send_register_sms") as send_sms:
            with self.captureOnCommitCallbacks(execute=True):
                patient = Patient.objects.create(
                    first_name="Ali",
                    last_name="Ahmadi",
                    mobile="09123456789",
                )

        send_sms.assert_called_once_with("09123456789", "Ali_Ahmadi")
        self.assertEqual(SMSMessageLog.objects.count(), 1)
        sms_log = SMSMessageLog.objects.get()
        self.assertEqual(sms_log.patient, patient)
        self.assertEqual(sms_log.mobile, "09123456789")
        self.assertEqual(sms_log.template, "register-template")
        self.assertEqual(sms_log.token, "Ali_Ahmadi")
        self.assertEqual(sms_log.status, SMSMessageLog.STATUS_SUCCESS)

    def test_sms_message_log_preserves_history_when_patient_is_deleted(self):
        patient = Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            national_code="1111111111",
        )
        sms_log = SMSMessageLog.objects.create(
            patient=patient,
            mobile=patient.mobile,
            template="done-template",
            token="Ali_Ahmadi",
            status=SMSMessageLog.STATUS_SUCCESS,
        )

        patient.delete()
        sms_log.refresh_from_db()

        self.assertIsNone(sms_log.patient)
        self.assertEqual(sms_log.mobile, "09123456789")

    @override_settings(
        KAVENEGAR_API_KEY="test-api-key",
        KAVENEGAR_DONE_TEMPLATE="done-template",
        SMS_SEND_ASYNC=False,
    )
    def test_admin_action_sends_done_sms_to_selected_patients(self):
        user = get_user_model().objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="password",
        )
        self.client.force_login(user)
        patient = Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            national_code="1111111111",
        )

        with patch("patients.sms_tasks.send_done_sms") as send_sms:
            response = self.client.post(
                reverse("admin:patients_patient_changelist"),
                {
                    "action": "send_done_sms_to_patients",
                    "_selected_action": [str(patient.pk)],
                    "index": "0",
                },
                follow=True,
            )

        self.assertEqual(response.status_code, 200)
        send_sms.assert_called_once_with("09123456789", "Ali_Ahmadi")
        sms_log = SMSMessageLog.objects.get(patient=patient)
        self.assertEqual(sms_log.template, "done-template")
        self.assertEqual(sms_log.token, "Ali_Ahmadi")
        self.assertEqual(sms_log.status, SMSMessageLog.STATUS_SUCCESS)
        patient.refresh_from_db()
        self.assertTrue(patient.done_sms_sent)
        self.assertIn(
            "ارسال پیامک انجام شد برای 1 بیمار در پس‌زمینه شروع شد.",
            [message.message for message in get_messages(response.wsgi_request)],
        )

    @override_settings(
        KAVENEGAR_API_KEY="test-api-key",
        KAVENEGAR_DONE_TEMPLATE="done-template",
        SMS_SEND_ASYNC=True,
    )
    def test_admin_action_queues_done_sms_without_waiting_for_delivery(self):
        user = get_user_model().objects.create_superuser(
            username="async-admin",
            email="async@example.com",
            password="password",
        )
        self.client.force_login(user)
        patient = Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            national_code="1111111111",
        )

        with patch("patients.admin.enqueue_done_sms_for_patients") as enqueue_done_sms:
            response = self.client.post(
                reverse("admin:patients_patient_changelist"),
                {
                    "action": "send_done_sms_to_patients",
                    "_selected_action": [str(patient.pk)],
                    "index": "0",
                },
                follow=True,
            )

        self.assertEqual(response.status_code, 200)
        enqueue_done_sms.assert_called_once_with([patient.pk])
        self.assertFalse(SMSMessageLog.objects.filter(patient=patient).exists())

    @override_settings(KAVENEGAR_API_KEY="", KAVENEGAR_DONE_TEMPLATE="")
    def test_admin_action_validates_kavenegar_settings_before_patient_loop(self):
        user = get_user_model().objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="password",
        )
        self.client.force_login(user)
        patient = Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            national_code="1111111111",
        )

        with patch("patients.admin.enqueue_done_sms_for_patients") as enqueue_done_sms:
            response = self.client.post(
                reverse("admin:patients_patient_changelist"),
                {
                    "action": "send_done_sms_to_patients",
                    "_selected_action": [str(patient.pk)],
                    "index": "0",
                },
                follow=True,
            )

        self.assertEqual(response.status_code, 200)
        enqueue_done_sms.assert_not_called()
        self.assertFalse(SMSMessageLog.objects.exists())
        self.assertIn(
            "تنظیمات سامانه پیامک (KAVENEGAR_API_KEY یا "
            "KAVENEGAR_DONE_TEMPLATE) پیکربندی نشده است.",
            [message.message for message in get_messages(response.wsgi_request)],
        )

    def test_admin_action_downloads_pdf_report_for_selected_patients(self):
        register_test_pdf_fonts()
        user = get_user_model().objects.create_superuser(
            username="admin-pdf",
            email="admin-pdf@example.com",
            password="password",
        )
        self.client.force_login(user)
        patient = Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            national_code="1111111111",
        )

        response = self.client.post(
            reverse("admin:patients_patient_changelist"),
            {
                "action": "download_patients_pdf_report",
                "_selected_action": [str(patient.pk)],
                "index": "0",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("selected-patients-report.pdf", response["Content-Disposition"])
        content = b"".join(response.streaming_content)
        self.assertTrue(content.startswith(b"%PDF"))

    def test_admin_action_downloads_excel_report_for_selected_patients(self):
        user = get_user_model().objects.create_superuser(
            username="admin-excel",
            email="admin-excel@example.com",
            password="password",
        )
        self.client.force_login(user)
        patient = Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            national_code="1111111111",
        )

        response = self.client.post(
            reverse("admin:patients_patient_changelist"),
            {
                "action": "download_patients_excel_report",
                "_selected_action": [str(patient.pk)],
                "index": "0",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("selected-patients-report.xlsx", response["Content-Disposition"])
        content = b"".join(response.streaming_content)
        workbook = load_workbook(BytesIO(content))
        worksheet = workbook.active
        self.assertEqual(
            [cell.value for cell in worksheet[1]],
            ["ردیف", "نام", "نام خانوادگی", "موبایل", "کد ملی", "زمان ثبت"],
        )
        self.assertEqual(
            [cell.value for cell in worksheet[2]],
            [
                1,
                "Ali",
                "Ahmadi",
                "09123456789",
                "1111111111",
                format_tehran_jalali(patient.created_at),
            ],
        )

    def test_build_patients_pdf_returns_pdf_buffer(self):
        register_test_pdf_fonts()
        patient = Patient.objects.create(
            first_name="علی",
            last_name="احمدی",
            mobile="09123456789",
            national_code="1111111111",
        )

        pdf_buffer = build_patients_pdf([patient])

        self.assertTrue(pdf_buffer.getvalue().startswith(b"%PDF"))

    def test_build_patients_excel_returns_xlsx_buffer(self):
        patient = Patient.objects.create(
            first_name="علی",
            last_name="احمدی",
            mobile="09123456789",
            national_code="1111111111",
        )

        excel_buffer = build_patients_excel([patient])
        workbook = load_workbook(excel_buffer)

        self.assertEqual(workbook.active["B2"].value, "علی")

    def test_patient_admin_list_shows_national_code_instead_of_mobile(self):
        model_admin = PatientAdmin(Patient, admin.site)

        self.assertEqual(
            model_admin.list_display,
            (
                "first_name",
                "last_name",
                "national_code",
                "done_sms_sent",
                "created_at_jalali",
            ),
        )
        self.assertEqual(model_admin.list_editable, ("done_sms_sent",))
        self.assertNotIn("mobile", model_admin.list_display)
        self.assertIn("mobile", model_admin.get_fields(request=Mock()))

    def test_patient_admin_national_code_field_has_copy_button_assets(self):
        form = PatientAdminForm()
        national_code_field = form.fields["national_code"]
        model_admin = PatientAdmin(Patient, admin.site)
        media = str(model_admin.media)

        self.assertEqual(
            national_code_field.widget.attrs["data-copy-national-code"], "true"
        )
        self.assertEqual(national_code_field.widget.attrs["maxlength"], "10")
        self.assertIn("copy_national_code.css", media)
        self.assertIn("copy_national_code.js", media)

    @override_settings(KAVENEGAR_DONE_TEMPLATE="done-template")
    def test_patient_admin_marks_patients_with_successful_done_sms(self):
        user = get_user_model().objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="password",
        )
        patient = Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            national_code="1111111111",
        )
        SMSMessageLog.objects.create(
            patient=patient,
            mobile=patient.mobile,
            template="done-template",
            token="Ali_Ahmadi",
            status=SMSMessageLog.STATUS_SUCCESS,
        )
        request = Mock(user=user)
        model_admin = PatientAdmin(Patient, admin.site)

        patient_from_admin = model_admin.get_queryset(request).get(pk=patient.pk)

        self.assertTrue(model_admin.sms_sent_indicator(patient_from_admin))

    @override_settings(KAVENEGAR_DONE_TEMPLATE="done-template")
    def test_patient_admin_does_not_mark_patients_without_successful_done_sms(self):
        user = get_user_model().objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="password",
        )
        patient = Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            national_code="1111111111",
        )
        SMSMessageLog.objects.create(
            patient=patient,
            mobile=patient.mobile,
            template="done-template",
            token="Ali_Ahmadi",
            status=SMSMessageLog.STATUS_FAILED,
        )
        request = Mock(user=user)
        model_admin = PatientAdmin(Patient, admin.site)

        patient_from_admin = model_admin.get_queryset(request).get(pk=patient.pk)

        self.assertFalse(model_admin.sms_sent_indicator(patient_from_admin))

    def test_patient_admin_manual_done_sms_tick_marks_patient(self):
        user = get_user_model().objects.create_superuser(
            username="manual-admin",
            email="manual@example.com",
            password="password",
        )
        patient = Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            national_code="1111111111",
        )
        request = Mock(user=user)
        model_admin = PatientAdmin(Patient, admin.site)

        Patient.objects.filter(pk=patient.pk).update(done_sms_sent=True)
        patient_from_admin = model_admin.get_queryset(request).get(pk=patient.pk)

        self.assertTrue(model_admin.sms_sent_indicator(patient_from_admin))

    @override_settings(KAVENEGAR_DONE_TEMPLATE="done-template")
    def test_patient_admin_ignores_successful_non_done_sms(self):
        user = get_user_model().objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="password",
        )
        patient = Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            national_code="1111111111",
        )
        SMSMessageLog.objects.create(
            patient=patient,
            mobile=patient.mobile,
            template="register-template",
            token="Ali_Ahmadi",
            status=SMSMessageLog.STATUS_SUCCESS,
        )
        request = Mock(user=user)
        model_admin = PatientAdmin(Patient, admin.site)

        patient_from_admin = model_admin.get_queryset(request).get(pk=patient.pk)

        self.assertFalse(model_admin.sms_sent_indicator(patient_from_admin))

    def test_admin_created_at_uses_jalali_date_and_tehran_time(self):
        model_admin = PatientAdmin(Patient, admin.site)
        patient = Patient(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            created_at=datetime(2026, 3, 20, 20, 45, 10, tzinfo=datetime_timezone.utc),
        )

        self.assertEqual(model_admin.created_at_jalali(patient), "۱۴۰۵/۰۱/۰۱ ۰۰:۱۵:۱۰")

    def test_sms_log_admin_created_at_uses_jalali_date_and_tehran_time(self):
        model_admin = SMSMessageLogAdmin(SMSMessageLog, admin.site)
        sms_log = SMSMessageLog(
            mobile="09123456789",
            template="done-template",
            token="Ali_Ahmadi",
            status=SMSMessageLog.STATUS_SUCCESS,
            created_at=datetime(2026, 6, 15, 8, 30, 5, tzinfo=datetime_timezone.utc),
        )

        self.assertEqual(model_admin.created_at_jalali(sms_log), "۱۴۰۵/۰۳/۲۵ ۱۲:۰۰:۰۵")

    def test_sms_response_is_formatted_for_admin_display(self):
        response = repr(
            [
                {
                    "messageid": 1540699584,
                    "message": "درخواست شما\nثبت شد",
                    "status": 5,
                    "statustext": "ارسال به مخابرات",
                    "sender": "10004347",
                    "receptor": "09164521571",
                    "date": 1781688729,
                    "cost": 2910,
                }
            ]
        )

        formatted_response = str(format_sms_response(response))

        self.assertIn("شناسه پیامک", formatted_response)
        self.assertIn("1540699584", formatted_response)
        self.assertIn("متن پیام", formatted_response)
        self.assertIn("درخواست شما\nثبت شد", formatted_response)
        self.assertIn("وضعیت سرویس", formatted_response)
        self.assertIn("ارسال به مخابرات", formatted_response)
        self.assertIn("زمان سرویس", formatted_response)
        self.assertIn("۱۴۰۵/۰۳/۲۷ ۱۳:۰۲:۰۹", formatted_response)
        self.assertIn("min-width: min(100%, 360px)", formatted_response)
        self.assertIn("overflow-wrap: break-word", formatted_response)
        self.assertIn("word-break: normal", formatted_response)
        self.assertNotIn("overflow-wrap: anywhere", formatted_response)
        self.assertNotIn("messageid", formatted_response)

    def test_sms_response_preserves_unparseable_text(self):
        formatted_response = str(format_sms_response("raw error response"))

        self.assertIn("raw error response", formatted_response)

    def test_sms_message_log_admin_views_do_not_allow_add_or_delete(self):
        user = get_user_model().objects.create_superuser(
            username="admin",
            email="admin@example.com",
            password="password",
        )
        request = Mock(user=user)
        inline = SMSMessageLogInline(Patient, admin.site)
        model_admin = SMSMessageLogAdmin(SMSMessageLog, admin.site)

        self.assertFalse(inline.has_add_permission(request))
        self.assertFalse(inline.has_delete_permission(request))
        self.assertFalse(model_admin.has_add_permission(request))
        self.assertFalse(model_admin.has_delete_permission(request))

    def test_sms_message_log_admin_hides_raw_response_fields(self):
        request = Mock(user=Mock())
        model_admin = SMSMessageLogAdmin(SMSMessageLog, admin.site)

        fields = model_admin.get_fields(request)

        self.assertIn("formatted_response", fields)
        self.assertIn("formatted_error", fields)
        self.assertNotIn("response", fields)
        self.assertNotIn("error", fields)

    @override_settings(KAVENEGAR_API_KEY="test-api-key")
    def test_patient_update_signal_does_not_send_register_sms(self):
        patient = Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            mobile="09123456789",
            national_code="1111111111",
        )

        with patch("patients.sms_tasks.send_register_sms") as send_sms:
            with self.captureOnCommitCallbacks(execute=True):
                patient.first_name = "Reza"
                patient.save()

        send_sms.assert_not_called()


class RegisterPatientViewTests(TestCase):
    def test_get_register_patient_displays_empty_form(self):
        response = self.client.get(reverse("patients:register"))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "patients/register.html")
        self.assertIsInstance(response.context["form"], PatientRegistrationForm)
        self.assertFalse(response.context["form"].is_bound)

    def test_register_template_includes_animated_registration_stats(self):
        Patient.objects.create(
            first_name="Ali",
            last_name="Ahmadi",
            national_code="1234567890",
            mobile="09123456789",
        )
        Patient.objects.create(
            first_name="Reza",
            last_name="Karimi",
            national_code="1234567891",
            mobile="09123456780",
        )

        response = self.client.get(reverse("patients:register"))

        self.assertEqual(
            response.context["stats"]["community_count"], COMMUNITY_BASE_COUNT + 2
        )
        self.assertEqual(response.context["stats"]["today_count"], 2)
        self.assertContains(
            response, f'data-counter data-target="{COMMUNITY_BASE_COUNT + 2}"'
        )
        self.assertContains(response, "افراد ثبت‌نام‌شده")
        self.assertContains(response, 'data-target="2"')
        self.assertNotContains(response, "ثبت‌نام‌شده در سایت")
        self.assertNotContains(response, "جامعه همراه طرح")
        self.assertContains(response, "چرا الان؟")
        self.assertContains(response, "مزیت‌های واقعی")
        self.assertContains(response, "تعرفه دولتی")

    def test_register_template_uses_persian_labels_and_submit_text(self):
        response = self.client.get(reverse("patients:register"))

        self.assertContains(response, SITE_NAME)
        self.assertContains(response, "نام")
        self.assertContains(response, "نام خانوادگی")
        self.assertContains(response, "شماره موبایل")
        self.assertContains(response, "ثبت اطلاعات و دریافت پیگیری درمانگاه")
        self.assertContains(response, "۰ از ۴ بخش تکمیل شده")
        self.assertContains(response, "data-sticky-cta")

    def test_register_template_includes_standard_copyright_footer(self):
        response = self.client.get(reverse("patients:register"))

        self.assertContains(response, '<footer class="site-footer')
        self.assertContains(response, "&copy; 2026 Helssa. All rights reserved.")

    def test_register_template_includes_bottom_social_contact_bar(self):
        response = self.client.get(reverse("patients:register"))

        self.assertContains(response, "سوالی داری؟")
        self.assertContains(response, 'href="https://ble.ir/helssaaa"')
        self.assertContains(response, 'href="https://eitaa.ir/helssaaa"')
        self.assertContains(response, 'aria-label="پرسش از هلسا در پیام‌رسان بله"')
        self.assertContains(response, 'aria-label="پرسش از هلسا در پیام‌رسان ایتا"')

    def test_register_template_includes_share_preview_metadata(self):
        response = self.client.get(reverse("patients:register"))

        share_image_url = f"http://testserver/static/{SHARE_IMAGE_PATH}"
        self.assertContains(response, f"<title>{SHARE_TITLE}</title>")
        self.assertContains(
            response, f'<meta name="description" content="{SHARE_DESCRIPTION}">'
        )
        self.assertContains(
            response, f'<meta property="og:site_name" content="{SITE_NAME}">'
        )
        self.assertContains(
            response, f'<meta property="og:title" content="{SHARE_TITLE}">'
        )
        self.assertContains(
            response,
            f'<meta property="og:description" content="{SHARE_DESCRIPTION}">',
        )
        self.assertContains(
            response, f'<meta property="og:url" content="{CANONICAL_URL}">'
        )
        self.assertContains(response, f'<link rel="canonical" href="{CANONICAL_URL}">')
        self.assertContains(
            response, f'<meta property="og:image" content="{share_image_url}">'
        )
        self.assertContains(
            response, '<meta property="og:image:type" content="image/png">'
        )
        self.assertContains(
            response, '<meta name="twitter:card" content="summary_large_image">'
        )
        self.assertContains(
            response, f'<meta name="twitter:image" content="{share_image_url}">'
        )
        self.assertContains(response, '<script type="application/ld+json">')
        self.assertContains(response, '"@type": "MedicalClinic"')
        self.assertContains(
            response, "درمانگاه ولیعصر صغاد - پزشک خانواده دکتر حسین شبانی"
        )

    def test_order_redirects_temporarily_to_medogram(self):
        response = self.client.get("/order/")

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response["Location"], "https://medogram.ir")

    def test_order_without_trailing_slash_uses_append_slash_redirect(self):
        response = self.client.get("/order")

        self.assertEqual(response.status_code, 301)
        self.assertEqual(response["Location"], "/order/")

    def test_register_alias_redirects_permanently_to_canonical_home(self):
        response = self.client.get(reverse("patients:register_patient"))

        self.assertEqual(response.status_code, 301)
        self.assertEqual(response["Location"], "/")

    def test_register_alias_redirect_preserves_query_string(self):
        response = self.client.get(
            reverse("patients:register_patient"),
            {"utm_source": "telegram", "utm_campaign": "family-doctor"},
        )

        self.assertEqual(response.status_code, 301)
        self.assertEqual(
            response["Location"],
            "/?utm_source=telegram&utm_campaign=family-doctor",
        )

    def test_robots_txt_allows_site_and_points_to_sitemap(self):
        response = self.client.get(reverse("patients:robots_txt"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/plain; charset=utf-8")
        content = response.content.decode()
        self.assertIn("User-agent: *", content)
        self.assertIn("Allow: /", content)
        self.assertIn(f"Sitemap: {CANONICAL_URL}sitemap.xml", content)

    def test_sitemap_xml_contains_only_canonical_homepage(self):
        response = self.client.get(reverse("patients:sitemap_xml"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/xml; charset=utf-8")
        self.assertContains(response, f"<loc>{CANONICAL_URL}</loc>")
        self.assertContains(response, "<changefreq>weekly</changefreq>")
        self.assertContains(response, "<priority>1.0</priority>")
        self.assertNotContains(response, "/register/")

    def test_logo_instructions_document_required_image_files(self):
        instructions = Path("logo.md").read_text(encoding="utf-8")

        self.assertIn(SHARE_IMAGE_PATH, instructions)
        self.assertIn(SITE_LOGO_PATH, instructions)
        self.assertIn("1200×630", instructions)
        self.assertIn("512×512", instructions)
        self.assertIn(SITE_NAME, instructions)
        self.assertIn(SHARE_TITLE, instructions)
        self.assertIn(SHARE_DESCRIPTION, instructions)

    def test_missing_share_image_downgrades_twitter_card(self):
        share_image_path = Path("patients/static") / SHARE_IMAGE_PATH
        hidden_share_image_path = share_image_path.with_suffix(".hidden-for-test")
        share_image_path.rename(hidden_share_image_path)
        try:
            response = self.client.get(reverse("patients:register"))
        finally:
            hidden_share_image_path.rename(share_image_path)

        self.assertContains(response, '<meta name="twitter:card" content="summary">')
        self.assertNotContains(response, '<meta name="twitter:image"')
        self.assertNotContains(response, '<meta property="og:image"')

    def test_missing_site_logo_does_not_render_broken_image(self):
        logo_path = Path("patients/static") / SITE_LOGO_PATH
        hidden_logo_path = logo_path.with_suffix(".hidden-for-test")
        logo_path.rename(hidden_logo_path)
        try:
            response = self.client.get(reverse("patients:register"))
        finally:
            hidden_logo_path.rename(logo_path)

        self.assertNotContains(response, 'class="hero__logo"')
        self.assertNotContains(response, '<link rel="icon" type="image/png"')

    def test_site_logo_renders_after_file_is_provided(self):
        logo_path = Path("patients/static") / SITE_LOGO_PATH
        logo_path.parent.mkdir(parents=True, exist_ok=True)
        logo_existed = logo_path.exists()
        original_logo = logo_path.read_bytes() if logo_existed else None
        logo_path.write_bytes(b"\x89PNG\r\n\x1a\n")
        try:
            response = self.client.get(reverse("patients:register"))
        finally:
            if logo_existed:
                logo_path.write_bytes(original_logo)
            else:
                logo_path.unlink(missing_ok=True)

        site_logo_url = f"http://testserver/static/{SITE_LOGO_PATH}"
        self.assertContains(response, 'class="hero__logo reveal"')
        self.assertContains(response, f'src="{site_logo_url}"')
        self.assertContains(response, f'href="{site_logo_url}"')

    def test_register_form_disables_submit_button_after_submit_with_javascript(self):
        response = self.client.get(reverse("patients:register"))

        self.assertContains(response, "data-registration-form")
        self.assertContains(response, "data-submit-button")
        self.assertContains(response, 'data-submitting-text="در حال ثبت..."')
        self.assertContains(response, "submitButton.disabled = true")

    def test_register_button_has_disabled_styles(self):
        css = Path("patients/static/patients/css/style.css").read_text()

        self.assertIn(".form-card button:disabled", css)
        self.assertIn("cursor: not-allowed", css)
        self.assertIn("--color-button-disabled", css)

    def test_register_template_uses_decorative_inline_svg_icons(self):
        response = self.client.get(reverse("patients:register"))

        self.assertContains(response, 'class="icon form-field__icon"', count=4)
        self.assertContains(response, 'focusable="false"')
        self.assertContains(response, 'aria-hidden="true"')

    def test_register_styles_use_short_motion_and_reduced_motion_override(self):
        css = Path("patients/static/patients/css/style.css").read_text()

        self.assertIn("@keyframes form-card-enter", css)
        self.assertIn("animation: form-card-enter 320ms ease-out both", css)
        self.assertIn("transition: background 0.2s ease", css)
        self.assertIn("@media (prefers-reduced-motion: reduce)", css)
        self.assertIn("transition-duration: 1ms !important", css)
        self.assertIn("animation: none", css)

    def test_register_template_uses_rtl_persian_html_attributes(self):
        response = self.client.get(reverse("patients:register"))

        self.assertContains(response, '<html lang="fa" dir="rtl">')

    def test_register_form_uses_persian_placeholders_and_ltr_mobile(self):
        response = self.client.get(reverse("patients:register"))

        self.assertContains(response, 'autocomplete="given-name"')
        self.assertContains(response, 'placeholder="مثلاً علی"')
        self.assertContains(response, 'autocomplete="family-name"')
        self.assertContains(response, 'placeholder="مثلاً رضایی"')
        self.assertContains(response, 'autocomplete="off"')
        self.assertContains(response, 'aria-describedby="national-code-help"')
        self.assertContains(response, 'maxlength="10"')
        self.assertContains(response, 'placeholder="1234567890"')
        self.assertContains(response, 'autocomplete="tel"')
        self.assertContains(response, 'aria-describedby="mobile-help"')
        self.assertContains(response, 'dir="ltr"')
        self.assertContains(response, 'inputmode="numeric"')
        self.assertContains(response, 'maxlength="11"')
        self.assertContains(response, 'placeholder="09123456789"')
        self.assertContains(response, "۱۱ رقم و شروع با 09.")

    def test_register_template_styles_messages_as_alert_cards(self):
        response = self.client.get(reverse("patients:register"))

        self.assertContains(response, 'class="message-stack"', count=0)

        response = self.client.post(
            reverse("patients:register"),
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "09123456789",
            },
            follow=True,
        )

        self.assertContains(response, 'class="message-stack"')
        self.assertContains(response, 'class="feedback-modal"')
        self.assertContains(response, 'class="feedback-dialog"')
        self.assertContains(response, 'role="dialog"')
        self.assertContains(response, 'aria-modal="true"')
        self.assertContains(response, "message-card message-card--success")
        self.assertContains(response, 'class="message-card__icon"')
        self.assertContains(response, "data-dismiss-feedback")
        self.assertContains(response, 'class="feedback-dialog__button"')
        self.assertContains(response, "باشه")
        self.assertContains(response, 'class="icon icon--status"')
        self.assertContains(response, 'aria-hidden="true"')
        self.assertNotContains(response, "✓")

    def test_field_errors_render_below_each_field(self):
        response = self.client.post(
            reverse("patients:register"),
            data={
                "first_name": "",
                "last_name": "",
                "national_code": "",
                "mobile": "08123456789",
            },
        )

        self.assertContains(response, 'aria-label="خطاهای نام"')
        self.assertContains(response, "نام را وارد کنید.")
        self.assertContains(response, 'aria-label="خطاهای نام خانوادگی"')
        self.assertContains(response, "نام خانوادگی را وارد کنید.")
        self.assertContains(response, 'aria-label="خطاهای شماره موبایل"')
        self.assertContains(response, "شماره موبایل با 09 شروع شود.")

    def test_invalid_post_preserves_submitted_values(self):
        response = self.client.post(
            reverse("patients:register"),
            data={
                "first_name": "علی",
                "last_name": "احمدی",
                "national_code": "1234567890",
                "mobile": "08123456789",
            },
        )

        self.assertContains(response, 'value="علی"')
        self.assertContains(response, 'value="احمدی"')
        self.assertContains(response, 'value="08123456789"')

    def test_post_valid_form_saves_patient_redirects_and_adds_success_message(self):
        response = self.client.post(
            reverse("patients:register"),
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "09123456789",
            },
        )

        self.assertRedirects(response, reverse("patients:register"))
        self.assertTrue(
            Patient.objects.filter(
                first_name="Ali",
                last_name="Ahmadi",
                mobile="09123456789",
            ).exists()
        )

        messages = list(response.wsgi_request._messages)
        self.assertEqual(len(messages), 1)
        self.assertEqual(str(messages[0]), SUCCESS_MESSAGE)

    def test_post_invalid_form_renders_errors_without_saving_patient(self):
        response = self.client.post(
            reverse("patients:register"),
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "08123456789",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Patient.objects.exists())
        self.assertContains(response, FORM_ERROR_MESSAGE)
        self.assertContains(response, 'class="message-card message-card--error"')
        self.assertContains(response, "باشه")
        self.assertContains(response, "شماره موبایل با 09 شروع شود.")
        self.assertTrue(response.context["form"].is_bound)

    def test_post_handles_duplicate_mobile_integrity_error(self):
        with patch.object(
            PatientRegistrationForm,
            "save",
            side_effect=IntegrityError(
                "UNIQUE constraint failed: patients_patient.mobile"
            ),
        ):
            response = self.client.post(
                reverse("patients:register"),
                data={
                    "first_name": "Ali",
                    "last_name": "Ahmadi",
                    "national_code": "1234567890",
                    "mobile": "09123456789",
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Patient.objects.exists())
        self.assertContains(response, FORM_ERROR_MESSAGE)
        self.assertContains(response, DUPLICATE_MOBILE_ERROR)
        self.assertEqual(
            response.context["form"].errors["mobile"], [DUPLICATE_MOBILE_ERROR]
        )

    def test_post_handles_generic_database_save_error(self):
        with patch.object(
            PatientRegistrationForm,
            "save",
            side_effect=DatabaseError("database unavailable"),
        ):
            response = self.client.post(
                reverse("patients:register"),
                data={
                    "first_name": "Ali",
                    "last_name": "Ahmadi",
                    "national_code": "1234567890",
                    "mobile": "09123456789",
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Patient.objects.exists())
        self.assertContains(response, "باشه")
        self.assertContains(
            response, "در ذخیره‌سازی اطلاعات مشکلی رخ داد. لطفاً دوباره تلاش کنید."
        )


class VisitAnalyticsTests(TestCase):
    def test_get_register_creates_form_view_event(self):
        from .models import VisitEvent

        response = self.client.get(reverse("patients:register"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            VisitEvent.objects.filter(
                event_type=VisitEvent.EVENT_FORM_VIEW, path="/"
            ).count(),
            1,
        )
        self.assertIn("helssa_vid", response.cookies)

    def test_get_register_alias_redirect_does_not_create_duplicate_visit_event(self):
        from .models import VisitEvent

        response = self.client.get(reverse("patients:register_patient"))

        self.assertEqual(response.status_code, 301)
        self.assertEqual(VisitEvent.objects.count(), 0)

    def test_repeated_register_get_from_same_visitor_is_deduplicated(self):
        first_response = self.client.get(reverse("patients:register"))
        second_response = self.client.get(reverse("patients:register"))

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(second_response.status_code, 200)
        self.assertEqual(
            VisitEvent.objects.filter(event_type=VisitEvent.EVENT_FORM_VIEW).count(),
            1,
        )

    @override_settings(ANALYTICS_PAGE_VIEW_DEDUP_SECONDS=0)
    def test_page_view_deduplication_can_be_disabled(self):
        self.client.get(reverse("patients:register"))
        self.client.get(reverse("patients:register"))

        self.assertEqual(
            VisitEvent.objects.filter(event_type=VisitEvent.EVENT_FORM_VIEW).count(),
            2,
        )

    def test_invalid_post_creates_attempt_and_invalid_events_without_values(self):
        from .models import VisitEvent

        self.client.post(
            reverse("patients:register"),
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "08123456789",
            },
        )

        self.assertTrue(
            VisitEvent.objects.filter(
                event_type=VisitEvent.EVENT_FORM_SUBMIT_ATTEMPT
            ).exists()
        )
        invalid = VisitEvent.objects.get(
            event_type=VisitEvent.EVENT_FORM_SUBMIT_INVALID
        )
        self.assertEqual(invalid.metadata, {"error_fields": ["mobile"]})
        self.assertNotIn("08123456789", str(invalid.metadata))

    def test_valid_post_creates_attempt_and_success_events(self):
        from .models import VisitEvent

        self.client.post(
            reverse("patients:register"),
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "09123456789",
            },
        )

        self.assertTrue(
            VisitEvent.objects.filter(
                event_type=VisitEvent.EVENT_FORM_SUBMIT_ATTEMPT
            ).exists()
        )
        success = VisitEvent.objects.get(
            event_type=VisitEvent.EVENT_FORM_SUBMIT_SUCCESS
        )
        self.assertEqual(success.patient.mobile, "09123456789")

    def test_logging_failure_never_crashes_registration(self):
        with patch("patients.views.log_visit_event", side_effect=Exception("boom")):
            response = self.client.post(
                reverse("patients:register"),
                data={
                    "first_name": "Ali",
                    "last_name": "Ahmadi",
                    "national_code": "1234567890",
                    "mobile": "09123456789",
                },
            )

        self.assertRedirects(response, reverse("patients:register"))
        self.assertTrue(Patient.objects.exists())

    def test_valid_post_sets_generated_visitor_cookie_on_redirect(self):
        response = self.client.post(
            reverse("patients:register"),
            data={
                "first_name": "Ali",
                "last_name": "Ahmadi",
                "national_code": "1234567890",
                "mobile": "09123456789",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertIn("helssa_vid", response.cookies)

    def test_summary_counts_known_events(self):
        from .analytics import get_visit_report_summary
        from .models import VisitEvent

        VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000001",
            event_type=VisitEvent.EVENT_FORM_VIEW,
            method="GET",
            path="/",
            status_code=200,
        )
        VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000002",
            event_type=VisitEvent.EVENT_FORM_SUBMIT_ATTEMPT,
            method="POST",
            path="/register/",
            referrer="https://example.com",
        )
        VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000002",
            event_type=VisitEvent.EVENT_FORM_SUBMIT_SUCCESS,
            method="POST",
            path="/register/",
            referrer="https://example.com",
        )

        summary = get_visit_report_summary(VisitEvent.objects.all())

        self.assertEqual(summary["total_events"], 3)
        self.assertEqual(summary["page_views"], 1)
        self.assertEqual(summary["unique_visitors"], 2)
        self.assertEqual(summary["form_views"], 1)
        self.assertEqual(summary["submit_attempts"], 1)
        self.assertEqual(summary["successful_registrations"], 1)
        self.assertEqual(summary["top_paths"][0]["path"], "/register/")
        self.assertEqual(summary["top_referrers"][0]["referrer"], "https://example.com")

    def test_visit_pdf_builder_returns_non_empty_pdf(self):
        from .admin import build_visit_events_pdf
        from .analytics import get_visit_report_summary
        from .models import VisitEvent
        from django.utils import timezone

        register_test_pdf_fonts()
        event = VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000001",
            event_type=VisitEvent.EVENT_FORM_VIEW,
            method="GET",
            path="/",
            status_code=200,
        )
        VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000002",
            event_type=VisitEvent.EVENT_APK_DOWNLOAD,
            method="GET",
            path="/down/helssa.apk",
            status_code=200,
        )
        queryset = VisitEvent.objects.all()
        summary = get_visit_report_summary(queryset)
        with patch(
            "patients.admin._rtl_text", side_effect=lambda value: str(value)
        ) as rtl_text:
            pdf = build_visit_events_pdf(
                queryset,
                summary,
                event.created_at,
                timezone.now(),
            )

        self.assertEqual(summary["apk_downloads"], 1)
        rtl_text.assert_any_call("دانلود اپلیکیشن")
        rtl_text.assert_any_call(1)
        self.assertGreater(len(pdf.getvalue()), 100)
        self.assertTrue(pdf.getvalue().startswith(b"%PDF"))

    def test_admin_report_does_not_expose_raw_editable_details(self):
        from .models import VisitEvent, VisitReport
        from .admin import VisitReportAdmin

        model_admin = VisitReportAdmin(VisitReport, admin.site)
        request = Mock(user=Mock())

        self.assertFalse(model_admin.has_add_permission(request))
        self.assertFalse(model_admin.has_change_permission(request, VisitEvent()))
        self.assertFalse(model_admin.has_delete_permission(request, VisitEvent()))

    @patch("patients.admin.timezone.now")
    def test_visit_report_default_range_is_last_three_hours(self, mocked_now):
        from .admin import _parse_report_range
        from django.test import RequestFactory
        from django.utils import timezone

        mocked_now.return_value = datetime(
            2026, 3, 21, 8, 30, tzinfo=datetime_timezone.utc
        )
        request = RequestFactory().get("/admin/patients/visitreport/")
        request.user = Mock()

        start_dt, end_dt, start_value, end_value, selected_range = _parse_report_range(
            request
        )

        self.assertEqual(selected_range, "custom")
        self.assertEqual(end_dt, mocked_now.return_value)
        self.assertEqual(
            start_dt, mocked_now.return_value - timezone.timedelta(hours=3)
        )
        self.assertEqual(start_value, "۱۴۰۵/۰۱/۰۱ ۰۹:۰۰")
        self.assertEqual(end_value, "۱۴۰۵/۰۱/۰۱ ۱۲:۰۰")

    @patch("patients.admin.timezone.now")
    def test_visit_report_today_shortcut_uses_tehran_day(self, mocked_now):
        from .admin import _parse_report_range
        from django.test import RequestFactory

        mocked_now.return_value = datetime(
            2026, 3, 21, 8, 30, tzinfo=datetime_timezone.utc
        )
        request = RequestFactory().get(
            "/admin/patients/visitreport/", {"range": "today"}
        )
        request.user = Mock()

        start_dt, end_dt, _start_value, _end_value, selected_range = (
            _parse_report_range(request)
        )

        self.assertEqual(selected_range, "today")
        self.assertEqual(format_tehran_jalali(start_dt), "۱۴۰۵/۰۱/۰۱ ۰۰:۰۰:۰۰")
        self.assertEqual(end_dt, mocked_now.return_value)

    def test_visit_report_summary_daily_counts_are_jalali_tehran_dates(self):
        from .analytics import get_visit_report_summary
        from .models import VisitEvent

        event = VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000001",
            event_type=VisitEvent.EVENT_FORM_VIEW,
            method="GET",
            path="/",
            status_code=200,
        )
        VisitEvent.objects.filter(pk=event.pk).update(
            created_at=datetime(2026, 3, 20, 20, 45, tzinfo=datetime_timezone.utc)
        )

        summary = get_visit_report_summary(VisitEvent.objects.all())

        self.assertEqual(list(summary["daily_counts"].keys()), ["۱۴۰۵/۰۱/۰۱"])

    def test_visit_report_recent_events_have_jalali_tehran_datetime(self):
        from .analytics import get_visit_report_summary
        from .models import VisitEvent

        event = VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000001",
            event_type=VisitEvent.EVENT_PAGE_VIEW,
            method="GET",
            path="/",
            status_code=200,
        )
        VisitEvent.objects.filter(pk=event.pk).update(
            created_at=datetime(2026, 6, 26, 8, 48, 10, tzinfo=datetime_timezone.utc)
        )

        summary = get_visit_report_summary(VisitEvent.objects.all())

        self.assertEqual(
            summary["recent_events"][0].created_at_jalali,
            "۱۴۰۵/۰۴/۰۵ ۱۲:۱۸:۱۰",
        )

    def test_visit_report_template_contains_shortcuts_and_print_button(self):
        user = get_user_model().objects.create_superuser(
            "admin", "admin@example.com", "pass"
        )
        self.client.force_login(user)

        response = self.client.get(reverse("admin:patients_visitreport_changelist"))

        self.assertContains(response, "امروز")
        self.assertContains(response, "دیروز")
        self.assertContains(response, "یک هفته قبل")
        self.assertContains(response, "یک ماه قبل")
        self.assertContains(response, "گزارش کل")
        self.assertContains(response, "window.print()")
        self.assertContains(response, "بازه گزارش:")
        self.assertContains(response, 'type="text" name="start_date"')
        self.assertContains(response, "فرمت نمونه: ۱۴۰۵/۰۴/۰۵ ۰۹:۰۰")
        self.assertNotContains(response, 'type="datetime-local"')

    def test_visit_report_today_shortcut_renders_without_admin_lookup_errors(self):
        user = get_user_model().objects.create_superuser(
            "range-admin", "range@example.com", "pass"
        )
        self.client.force_login(user)

        response = self.client.get(
            reverse("admin:patients_visitreport_changelist"), {"range": "today"}
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'href="?range=today"')
        self.assertContains(response, 'href="export-pdf/?range=today"')


class VisitAnalyticsEnhancedTests(TestCase):
    def test_masked_ip_for_ipv4_masks_last_octet(self):
        from .analytics import mask_ip

        self.assertEqual(mask_ip("5.122.34.77"), "5.122.34.xxx")

    def test_client_engagement_endpoint_logs_allowed_event_without_sensitive_metadata(
        self,
    ):
        response = self.client.post(
            reverse("patients:analytics_event"),
            data=json.dumps(
                {
                    "event_type": VisitEvent.EVENT_FIELD_COMPLETE,
                    "metadata": {
                        "field_name": "mobile",
                        "mobile": "09123456789",
                        "national_code": "1234567890",
                        "section": "form",
                        "raw": {"mobile": "09123456789"},
                    },
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        event = VisitEvent.objects.get(event_type=VisitEvent.EVENT_FIELD_COMPLETE)
        self.assertEqual(event.metadata, {"section": "form", "field_name": "mobile"})
        self.assertNotIn("09123456789", str(event.metadata))
        self.assertNotIn("1234567890", str(event.metadata))

    def test_client_engagement_endpoint_rejects_unknown_event(self):
        response = self.client.post(
            reverse("patients:analytics_event"),
            data=json.dumps({"event_type": "not_allowed", "metadata": {}}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(VisitEvent.objects.exists())

    def test_register_template_wires_engagement_tracking(self):
        response = self.client.get(reverse("patients:register"))

        self.assertContains(response, reverse("patients:analytics_event"))
        self.assertContains(response, 'data-track-click="hero_cta_click"')
        self.assertContains(response, 'data-track-click="sticky_cta_click"')
        self.assertContains(response, 'data-track-section="faq"')
        self.assertContains(response, "sendEngagement")

    @override_settings(ANALYTICS_STORE_RAW_IP=False)
    def test_raw_ip_is_not_saved_when_disabled(self):
        from .models import VisitEvent

        self.client.get(reverse("patients:register"), REMOTE_ADDR="5.122.34.77")
        event = VisitEvent.objects.get()
        self.assertIsNone(event.ip_address)
        self.assertEqual(event.masked_ip, "5.122.34.xxx")

    @override_settings(ANALYTICS_STORE_RAW_IP=True)
    def test_raw_ip_is_saved_when_enabled(self):
        from .models import VisitEvent

        self.client.get(reverse("patients:register"), REMOTE_ADDR="5.122.34.77")
        self.assertEqual(str(VisitEvent.objects.get().ip_address), "5.122.34.77")

    def test_utm_source_and_campaign_are_saved_from_url(self):
        from .models import VisitEvent

        self.client.get(
            reverse("patients:register") + "?utm_source=telegram&utm_campaign=summer"
        )
        event = VisitEvent.objects.get()
        self.assertEqual(event.utm_source, "telegram")
        self.assertEqual(event.utm_campaign, "summer")

    def test_mobile_user_agent_sets_device_type_mobile(self):
        from .models import VisitEvent

        self.client.get(
            reverse("patients:register"),
            HTTP_USER_AGENT="Mozilla/5.0 (iPhone; CPU iPhone OS) Mobile Safari/604.1",
        )
        self.assertEqual(VisitEvent.objects.get().device_type, "mobile")

    def test_bot_user_agent_sets_is_bot_true(self):
        from .models import VisitEvent

        self.client.get(reverse("patients:register"), HTTP_USER_AGENT="Googlebot/2.1")
        event = VisitEvent.objects.get()
        self.assertTrue(event.is_bot)
        self.assertEqual(event.device_type, "bot")

    def test_summary_calculates_conversion_rate(self):
        from .analytics import get_visit_report_summary
        from .models import VisitEvent

        for index, event_type in enumerate(
            [
                VisitEvent.EVENT_FORM_VIEW,
                VisitEvent.EVENT_FORM_VIEW,
                VisitEvent.EVENT_FORM_SUBMIT_ATTEMPT,
                VisitEvent.EVENT_FORM_SUBMIT_SUCCESS,
            ]
        ):
            VisitEvent.objects.create(
                visitor_id=f"00000000-0000-0000-0000-00000000000{index+1}",
                event_type=event_type,
                method="GET",
                path="/",
            )
        summary = get_visit_report_summary(VisitEvent.objects.all())
        self.assertEqual(summary["conversion_rate"], 50.0)
        self.assertEqual(summary["submit_success_rate"], 100.0)

    def test_device_cards_count_unique_visitors_not_events(self):
        from .analytics import get_visit_report_summary

        for _ in range(3):
            VisitEvent.objects.create(
                visitor_id="00000000-0000-0000-0000-000000000001",
                event_type=VisitEvent.EVENT_FORM_VIEW,
                method="GET",
                path="/",
                device_type="mobile",
            )
        VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000002",
            event_type=VisitEvent.EVENT_FORM_VIEW,
            method="GET",
            path="/",
            device_type="desktop",
        )

        summary = get_visit_report_summary(VisitEvent.objects.all())

        self.assertEqual(summary["mobile_count"], 1)
        self.assertEqual(summary["desktop_count"], 1)

    @override_settings(ANALYTICS_SHOW_RAW_IP_TO_SUPERUSER=True)
    def test_admin_report_superuser_can_see_raw_ip_when_allowed(self):
        from .models import VisitEvent

        user = get_user_model().objects.create_superuser(
            "raw-admin", "raw@example.com", "pass"
        )
        self.client.force_login(user)
        VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000001",
            event_type=VisitEvent.EVENT_FORM_VIEW,
            method="GET",
            path="/",
            masked_ip="5.122.34.xxx",
            ip_address="5.122.34.77",
        )
        response = self.client.get(
            reverse("admin:patients_visitreport_changelist"), {"range": "all"}
        )
        self.assertContains(response, "5.122.34.77")

    @override_settings(ANALYTICS_SHOW_RAW_IP_TO_SUPERUSER=True)
    def test_admin_report_staff_cannot_see_raw_ip(self):
        from .models import VisitEvent

        user = get_user_model().objects.create_user(
            "staff", "staff@example.com", "pass", is_staff=True
        )
        user.user_permissions.add(*[])
        user.is_superuser = False
        user.save()
        self.client.force_login(user)
        VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000001",
            event_type=VisitEvent.EVENT_FORM_VIEW,
            method="GET",
            path="/",
            masked_ip="5.122.34.xxx",
            ip_address="5.122.34.77",
        )
        response = self.client.get(
            reverse("admin:patients_visitreport_changelist"), {"range": "all"}
        )
        self.assertNotContains(response, "5.122.34.77")
        self.assertContains(response, "5.122.34.xxx")

    def test_export_csv_respects_filters(self):
        from .models import VisitEvent

        user = get_user_model().objects.create_superuser(
            "csv-admin", "csv@example.com", "pass"
        )
        self.client.force_login(user)
        VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000001",
            event_type=VisitEvent.EVENT_FORM_VIEW,
            method="GET",
            path="/",
            device_type="mobile",
        )
        VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000002",
            event_type=VisitEvent.EVENT_FORM_VIEW,
            method="GET",
            path="/",
            device_type="desktop",
        )
        response = self.client.get(
            reverse("admin:patients_visitreport_export_csv"),
            {"range": "all", "device_type": "mobile"},
        )
        content = response.content.decode("utf-8")
        self.assertEqual(response.status_code, 200)
        self.assertIn("mobile", content)
        self.assertNotIn("desktop", content)

    def test_pdf_and_csv_links_preserve_active_filters(self):
        user = get_user_model().objects.create_superuser(
            "link-admin", "link@example.com", "pass"
        )
        self.client.force_login(user)
        response = self.client.get(
            reverse("admin:patients_visitreport_changelist"),
            {"range": "all", "device_type": "mobile", "utm_source": "telegram"},
        )
        self.assertContains(
            response,
            "export-pdf/?range=all&amp;device_type=mobile&amp;utm_source=telegram",
        )
        self.assertContains(
            response,
            "export-csv/?range=all&amp;device_type=mobile&amp;utm_source=telegram",
        )


class ApkDownloadTests(TestCase):
    def test_download_endpoint_serves_apk_and_logs_success(self):
        from pathlib import Path
        from tempfile import TemporaryDirectory
        from unittest.mock import patch

        from .models import VisitEvent

        with TemporaryDirectory() as tmpdir:
            apk_path = Path(tmpdir) / "helssa.apk"
            apk_path.write_bytes(b"fake-apk")
            with patch("patients.views.APK_DOWNLOAD_PATH", apk_path):
                response = self.client.get("/down/helssa.apk")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"], "application/vnd.android.package-archive"
        )
        self.assertIn("helssa.apk", response["Content-Disposition"])
        self.assertEqual(b"".join(response.streaming_content), b"fake-apk")
        event = VisitEvent.objects.get(event_type=VisitEvent.EVENT_APK_DOWNLOAD)
        self.assertEqual(event.path, "/down/helssa.apk")
        self.assertEqual(event.status_code, 200)
        self.assertEqual(event.metadata["result"], "download")

    def test_download_endpoint_logs_missing_file_as_404(self):
        from pathlib import Path
        from tempfile import TemporaryDirectory
        from unittest.mock import patch

        from .models import VisitEvent

        with TemporaryDirectory() as tmpdir, patch(
            "patients.views.APK_DOWNLOAD_PATH", Path(tmpdir) / "helssa.apk"
        ):
            response = self.client.get("/down/helssa.apk")

        self.assertEqual(response.status_code, 404)
        event = VisitEvent.objects.get(event_type=VisitEvent.EVENT_APK_DOWNLOAD)
        self.assertEqual(event.status_code, 404)
        self.assertEqual(event.metadata["result"], "missing")

    def test_download_endpoint_uses_actual_apk_filename_when_default_missing(self):
        from tempfile import TemporaryDirectory

        with TemporaryDirectory() as tmpdir:
            download_dir = Path(tmpdir)
            apk_path = download_dir / "helssa-release-v2.apk"
            apk_path.write_bytes(b"versioned-apk")
            with override_settings(
                APK_DOWNLOAD_PATH=download_dir / "helssa.apk",
                APK_DOWNLOAD_DIR=download_dir,
            ):
                response = self.client.get("/down/helssa.apk")

        self.assertEqual(response.status_code, 200)
        self.assertIn("helssa-release-v2.apk", response["Content-Disposition"])
        self.assertEqual(b"".join(response.streaming_content), b"versioned-apk")

    def test_admin_index_exposes_authenticated_apk_download_link(self):
        user = get_user_model().objects.create_superuser(
            username="apk-admin",
            email="apk-admin@example.com",
            password="password",
        )
        self.client.force_login(user)

        response = self.client.get(reverse("admin:index"))

        self.assertContains(response, "دانلود فایل APK")
        self.assertContains(response, reverse("admin_download_helssa_apk"))

    def test_admin_index_exposes_apk_upload_form(self):
        user = get_user_model().objects.create_superuser(
            username="apk-upload-admin",
            email="apk-upload-admin@example.com",
            password="password",
        )
        self.client.force_login(user)

        response = self.client.get(reverse("admin:index"))

        self.assertContains(response, "آپلود فایل APK جدید")
        self.assertContains(response, reverse("admin_upload_helssa_apk"))
        self.assertContains(response, 'name="apk_file"')

    def test_admin_upload_endpoint_saves_file_with_configured_download_name(self):
        from tempfile import TemporaryDirectory

        user = get_user_model().objects.create_superuser(
            username="apk-uploader",
            email="apk-uploader@example.com",
            password="password",
        )
        self.client.force_login(user)

        with TemporaryDirectory() as tmpdir:
            download_dir = Path(tmpdir) / "down"
            apk_path = download_dir / "helssa.apk"
            uploaded_file = SimpleUploadedFile(
                "release-v9.apk",
                b"uploaded-apk",
                content_type="application/vnd.android.package-archive",
            )
            with override_settings(
                APK_DOWNLOAD_PATH=apk_path,
                APK_DOWNLOAD_DIR=download_dir,
            ):
                response = self.client.post(
                    reverse("admin_upload_helssa_apk"),
                    {"apk_file": uploaded_file},
                    follow=True,
                )

                self.assertRedirects(response, reverse("admin:index"))
                self.assertTrue(apk_path.exists())
                self.assertEqual(apk_path.read_bytes(), b"uploaded-apk")
                self.assertFalse((download_dir / "release-v9.apk").exists())
                download_response = self.client.get(
                    reverse("admin_download_helssa_apk")
                )

        self.assertEqual(download_response.status_code, 200)
        self.assertIn("helssa.apk", download_response["Content-Disposition"])
        self.assertEqual(b"".join(download_response.streaming_content), b"uploaded-apk")

    def test_admin_upload_endpoint_rejects_non_apk_file(self):
        from tempfile import TemporaryDirectory

        user = get_user_model().objects.create_superuser(
            username="apk-upload-validator",
            email="apk-upload-validator@example.com",
            password="password",
        )
        self.client.force_login(user)

        with TemporaryDirectory() as tmpdir:
            download_dir = Path(tmpdir)
            apk_path = download_dir / "helssa.apk"
            uploaded_file = SimpleUploadedFile("notes.txt", b"not-apk")
            with override_settings(
                APK_DOWNLOAD_PATH=apk_path,
                APK_DOWNLOAD_DIR=download_dir,
            ):
                response = self.client.post(
                    reverse("admin_upload_helssa_apk"),
                    {"apk_file": uploaded_file},
                    follow=True,
                )

        self.assertRedirects(response, reverse("admin:index"))
        self.assertFalse(apk_path.exists())
        messages = [message.message for message in get_messages(response.wsgi_request)]
        self.assertIn("فقط فایل با پسوند APK قابل آپلود است.", messages)

    def test_admin_download_endpoint_serves_apk_with_actual_filename(self):
        from tempfile import TemporaryDirectory

        user = get_user_model().objects.create_superuser(
            username="apk-download-admin",
            email="apk-download-admin@example.com",
            password="password",
        )
        self.client.force_login(user)

        with TemporaryDirectory() as tmpdir:
            download_dir = Path(tmpdir)
            apk_path = download_dir / "helssa-admin-release.apk"
            apk_path.write_bytes(b"admin-apk")
            with override_settings(
                APK_DOWNLOAD_PATH=download_dir / "helssa.apk",
                APK_DOWNLOAD_DIR=download_dir,
            ):
                response = self.client.get(reverse("admin_download_helssa_apk"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"], "application/vnd.android.package-archive"
        )
        self.assertIn("helssa-admin-release.apk", response["Content-Disposition"])
        self.assertEqual(b"".join(response.streaming_content), b"admin-apk")
        self.assertFalse(
            VisitEvent.objects.filter(event_type=VisitEvent.EVENT_APK_DOWNLOAD).exists()
        )

    def test_qr_endpoint_returns_svg_for_download_url(self):
        response = self.client.get("/down/helssa-qr.svg")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "image/svg+xml; charset=utf-8")
        self.assertIn(b"<svg", response.content)

    def test_register_page_exposes_apk_download_link_and_qr_code(self):
        response = self.client.get("/")

        self.assertContains(response, 'href="http://testserver/down/helssa.apk"')
        self.assertContains(response, ">دانلود هلسا</a>")
        self.assertContains(response, 'src="http://testserver/down/helssa-qr.svg"')

    def test_summary_counts_apk_download_events(self):
        from .analytics import get_visit_report_summary
        from .models import VisitEvent

        VisitEvent.objects.create(
            visitor_id="00000000-0000-0000-0000-000000000001",
            event_type=VisitEvent.EVENT_APK_DOWNLOAD,
            method="GET",
            path="/down/helssa.apk",
            status_code=200,
        )

        summary = get_visit_report_summary(VisitEvent.objects.all())

        self.assertEqual(summary["apk_downloads"], 1)
